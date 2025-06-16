from flask import Flask, request, send_file, render_template_string
import pandas as pd
import openpyxl
from io import BytesIO
import datetime
import zipfile
import os

app = Flask(__name__)

HTML_TEMPLATE = """
<!doctype html>
<html lang=\"ko\">
  <head>
    <meta charset=\"utf-8\">
    <title>다품목 발주서 변환기</title>
  </head>
  <body style=\"font-family: sans-serif; text-align: center; margin-top: 50px;\">
    <h1>📦 다품목 발주서 변환기</h1>
    <form action=\"/convert\" method=\"post\" enctype=\"multipart/form-data\">
      <label>DeliveryList 파일 (.xlsx)</label><br>
      <input type=\"file\" name=\"delivery_file\" accept=\".xlsx\" required><br><br>
      <label>공통 양식 파일 (.xlsx)</label><br>
      <input type=\"file\" name=\"common_template\" accept=\".xlsx\" required><br><br>
      <label>의성복숭아 양식 파일 (.xlsx)</label><br>
      <input type=\"file\" name=\"uiseong_template\" accept=\".xlsx\" required><br><br>
      <button type=\"submit\">발주서 생성</button>
    </form>
  </body>
</html>
"""

@app.route("/", methods=["GET"])
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route("/convert", methods=["POST"])
def convert():
    delivery_file = request.files['delivery_file']
    common_template_file = request.files['common_template']
    uiseong_template_file = request.files['uiseong_template']

    df = pd.read_excel(delivery_file)
    # 컬럼명 공백 제거 처리
    df.columns = df.columns.str.replace(' ', '')

    required_columns = ["수취인이름", "구매자전화번호", "우편번호", "등록옵션명", "구매수(수량)", "등록상품명"]
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        return f"❌ DeliveryList에 필요한 컬럼이 없습니다: {', '.join(missing)}", 400

    df["수취인명"] = df["수취인이름"]
    df["수취인전화번호"] = df["구매자전화번호"]
    df["수취인이동통신"] = df["구매자전화번호"]
    df["수취인우편번호"] = df["우편번호"]
    df["주문상품명"] = df["등록옵션명"]
    df["상품모델"] = df["등록옵션명"]
    df["수량"] = df["구매수(수량)"]

    common_keywords = ["천도복숭아", "신비복숭아", "신틸라"]
    uiseong_keyword = "의성프리미엄신비복숭아"

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w') as zip_buffer:
        # 공통 품목
        common_df = df[df['등록상품명'].str.contains('|'.join(common_keywords), na=False)]
        if not common_df.empty:
            wb = openpyxl.load_workbook(common_template_file)
            ws = wb.active
            start_row = 2
            for _, row in common_df.iterrows():
                ws.cell(row=start_row, column=1).value = row['수취인명']
                ws.cell(row=start_row, column=2).value = row['수취인전화번호']
                ws.cell(row=start_row, column=3).value = row['수취인이동통신']
                ws.cell(row=start_row, column=4).value = row['수취인우편번호']
                ws.cell(row=start_row, column=5).value = row['주문상품명']
                ws.cell(row=start_row, column=6).value = row['상품모델']
                ws.cell(row=start_row, column=7).value = row['수량']
                start_row += 1
            temp = BytesIO()
            today = datetime.datetime.now().strftime('%y%m%d')
            wb.save(temp)
            temp.seek(0)
            zip_buffer.writestr(f"공통발주서_{today}.xlsx", temp.read())

        # 의성 품목
        uiseong_df = df[df['등록상품명'].str.contains(uiseong_keyword, na=False)]
        if not uiseong_df.empty:
            wb = openpyxl.load_workbook(uiseong_template_file)
            ws = wb.active
            start_row = 2
            for _, row in uiseong_df.iterrows():
                ws.cell(row=start_row, column=1).value = row['수취인명']
                ws.cell(row=start_row, column=2).value = row['수취인전화번호']
                ws.cell(row=start_row, column=3).value = row['수취인이동통신']
                ws.cell(row=start_row, column=4).value = row['수취인우편번호']
                ws.cell(row=start_row, column=5).value = row['주문상품명']
                ws.cell(row=start_row, column=6).value = row['상품모델']
                ws.cell(row=start_row, column=7).value = row['수량']
                start_row += 1
            temp = BytesIO()
            today = datetime.datetime.now().strftime('%y%m%d')
            wb.save(temp)
            temp.seek(0)
            zip_buffer.writestr(f"의성발주서_{today}.xlsx", temp.read())

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name="발주서_모음.zip",
                     mimetype="application/zip")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
